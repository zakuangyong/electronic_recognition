"""
M6: Generate diff report — summary.json and diff_report.xlsx.

Usage:
    python generate_report.py work/regions_all.json old.pdf new.pdf -o output/
"""

import argparse
import json
from pathlib import Path
from datetime import datetime


def build_summary(
    old_source: str,
    new_source: str,
    old_pdf: str,
    new_pdf: str,
    pages_data: list[dict],
) -> dict:
    """Build the summary.json structure per the plan spec."""
    total_regions = sum(len(p.get("regions", [])) for p in pages_data)
    text_regions = sum(
        1 for p in pages_data
        for r in p.get("regions", [])
        if r.get("old_text") or r.get("new_text")
    )

    summary = {
        "old_source": old_source,
        "new_source": new_source,
        "old_pdf": old_pdf,
        "new_pdf": new_pdf,
        "generated_at": datetime.now().isoformat(),
        "total_pages": len(pages_data),
        "total_regions": total_regions,
        "text_extracted_regions": text_regions,
        "pages": [],
    }

    for page_data in pages_data:
        page_entry = {
            "page": page_data["page"],
            "offset_px": page_data.get("offset_px", [0, 0]),
            "region_count": len(page_data.get("regions", [])),
            "regions": [],
        }

        for region in page_data.get("regions", []):
            old_text = region.get("old_text", "")
            new_text = region.get("new_text", "")
            if (old_text or new_text) and old_text != new_text:
                change_type = "text_changed"
            else:
                change_type = "visual_change"

            page_entry["regions"].append({
                "region_id": region["region_id"],
                "bbox_px": region["bbox_px"],
                "change_type": change_type,
                "old_text": old_text,
                "new_text": new_text,
                "old_crop": region.get("old_crop", ""),
                "new_crop": region.get("new_crop", ""),
                "confidence": region.get("confidence", 0.85),
                "review_status": "pending",
            })

        summary["pages"].append(page_entry)

    return summary


def write_excel_report(summary: dict, output_path: Path):
    """Write diff_report.xlsx from summary data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "差异清单"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["页码", "差异编号", "变更类型", "坐标(x0,y0,x1,y1)", "旧版文本", "新版文本",
               "旧版截图", "新版截图", "置信度", "复核状态"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    row = 2
    for page_data in summary["pages"]:
        for region in page_data["regions"]:
            bbox = region["bbox_px"]
            bbox_str = f"({bbox[0]},{bbox[1]},{bbox[2]},{bbox[3]})"

            values = [
                page_data["page"],
                region["region_id"],
                region["change_type"],
                bbox_str,
                region.get("old_text", ""),
                region.get("new_text", ""),
                region.get("old_crop", ""),
                region.get("new_crop", ""),
                region.get("confidence", ""),
                region.get("review_status", "pending"),
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if region.get("review_status") == "pending":
                    cell.fill = pending_fill

            row += 1

    col_widths = [8, 10, 30, 30, 25, 25, 50, 50, 10, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("汇总")
    ws2.cell(row=1, column=1, value="项目").font = header_font
    ws2.cell(row=1, column=2, value="内容").font = header_font

    info_rows = [
        ("旧版文件", summary["old_source"]),
        ("新版文件", summary["new_source"]),
        ("旧版PDF", summary["old_pdf"]),
        ("新版PDF", summary["new_pdf"]),
        ("生成时间", summary["generated_at"]),
        ("总页数", summary["total_pages"]),
        ("差异区域总数", summary["total_regions"]),
        ("文本提取区域数", summary["text_extracted_regions"]),
    ]
    for i, (key, val) in enumerate(info_rows, 2):
        ws2.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=str(val))
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 60

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Excel report saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Generate diff report")
    parser.add_argument("--old-source", type=str, default="raw.CATDrawing", help="Old source file name")
    parser.add_argument("--new-source", type=str, default="three_change.CATDrawing", help="New source file name")
    parser.add_argument("--old-pdf", type=str, default="raw.pdf", help="Old PDF file name")
    parser.add_argument("--new-pdf", type=str, default="three_change.pdf", help="New PDF file name")
    parser.add_argument("-i", "--input", type=Path, required=True, help="All-regions JSON file")
    parser.add_argument("-o", "--output", type=Path, required=True, help="Output directory")
    args = parser.parse_args()

    if not args.input.exists():
        print(f"Error: input file not found: {args.input}")
        return 1

    with open(args.input, "r", encoding="utf-8") as f:
        pages_data = json.load(f)

    summary = build_summary(
        args.old_source, args.new_source,
        args.old_pdf, args.new_pdf,
        pages_data,
    )

    summary_path = args.output / "summary.json"
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)
    print(f"Summary saved: {summary_path}")

    excel_path = args.output / "diff_report.xlsx"
    write_excel_report(summary, excel_path)

    print(f"Report generated: {len(summary['pages'])} page(s), {summary['total_regions']} region(s)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
